import psycopg2
import openpyxl

conn = psycopg2.connect(host='localhost', 
	database="DBLIST", user="postgres", password="pass")
cur = conn.cursor()

#student
cur.execute('''
DROP TABLE IF EXISTS Student;
CREATE TABLE IF NOT EXISTS Student (
    Num       VARCHAR(20),     --序号
    Sno       VARCHAR(20),    --学号
    Sname     TEXT,           --姓名
    Sclass    TEXT,           --班级
    PRIMARY KEY(Num));
''')
#course
cur.execute('''
DROP TABLE IF EXISTS Course;
CREATE TABLE IF NOT EXISTS Course  (
    Num         VARCHAR(20),     --序号
    Cno         VARCHAR(20),     --课程号
    Cname       TEXT,            --课程名称
    Cteacher    TEXT,            --任课老师
    Ccredit     VARCHAR(20),      --课程学分
    Cdate       TEXT,            --课程时间
    PRIMARY KEY(Num));
''')
#SC
cur.execute('''
DROP TABLE IF EXISTS SC;
CREATE TABLE IF NOT EXISTS SC (
    SC_Sno      VARCHAR(20),     --学生序号
    SC_Cno      VARCHAR(20),     --课程序号
    SC_Cname    TEXT,            --课程名称
    Grade       NUMERIC(5,2),    --成绩
    PRIMARY KEY(SC_Sno,SC_Cno));
''')


wb=openpyxl.load_workbook("list.xlsx")
ws=wb.active

#student
colC=ws["C"]
colD=ws["D"]
#course
colG=ws["G"]
colH=ws["H"]
colI=ws["I"]
colJ=ws["J"]
colK=ws["K"]
#SC
colM=ws["M"]
colN=ws["N"]
colO=ws["O"]
colP=ws["P"]



#student
for i in range(1,31):
    Num = i
    Sno=i
    Sname = '%s' % colC[i].value
    Sclass = '%s' % colD[i].value
    cur.execute('''
		INSERT INTO Student(Num,Sno,Sname,Sclass) VALUES (%(Num)s, %(Sno)s,%(Sname)s,%(Sclass)s) 
	''', {'Num':Num, 'Sno':Sno,'Sname':Sname,"Sclass":Sclass} )
#course
for i in range(1,9):
    Num = i
    Cno= '%s' % colG[i].value
    Cname = '%s' % colH[i].value
    Cteacher = '%s' % colI[i].value
    Ccredit='%s' % colJ[i].value
    Cdate = '%s' % colK[i].value
    cur.execute('''
		INSERT INTO Course(Num,Cno,Cname,Cteacher,Ccredit,Cdate) VALUES (%(Num)s, %(Cno)s,%(Cname)s,%(Cteacher)s,%(Ccredit)s,%(Cdate)s) 
	''', {'Num':Num, 'Cno':Cno,'Cname':Cname,"Cteacher":Cteacher,"Ccredit":Ccredit,"Cdate":Cdate} )
#SC
for i in range(1,len(colM)):
    SC_Sno = '%s' % colM[i].value
    SC_Cno= '%s' % colN[i].value
    SC_Cname='%s'%colO[i].value
    Grade='%s' % colP[i].value
    cur.execute('''
		INSERT INTO SC(SC_Sno,SC_Cno,SC_Cname,Grade) VALUES (%(SC_Sno)s,%(SC_Cno)s,%(SC_Cname)s,%(Grade)s) 
	''', {"SC_Sno":SC_Sno, "SC_Cno":SC_Cno,"SC_Cname":SC_Cname,"Grade":Grade} )



conn.commit()